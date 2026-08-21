import os
from pathlib import Path
from datetime import date

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")

import django
django.setup()

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from apps.cobranca.resources.devedor import (
    DevedoresResource,
    DevedoresEnderecosResource,
    DevedoresContatosResource,
    ContratosParcelasResource,
)

from apps.cobranca.models.carteira import Carteiras


OUTPUT_DIR = Path(__file__).resolve().parent / "arquivos_teste_importacao"

HEADER_FILL = "123B5D"
HEADER_FONT = "FFFFFF"


def get_import_headers(resource_class):
    """
    Get the import columns directly from the django-import-export Resource.
    """
    resource = resource_class()

    if hasattr(resource, "get_import_fields"):
        fields = resource.get_import_fields()
    else:
        fields = resource.fields.values()

    headers = []

    for field in fields:
        column_name = getattr(field, "column_name", None)

        if column_name and column_name not in headers:
            headers.append(column_name)

    return headers


def apply_header_style(ws):
    for cell in ws[1]:
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=HEADER_FILL,
        )
        cell.font = Font(
            bold=True,
            color=HEADER_FONT,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def adjust_column_widths(ws):
    for column_cells in ws.columns:
        max_length = 0

        for cell in column_cells:
            if cell.value is None:
                continue

            max_length = max(
                max_length,
                len(str(cell.value)),
            )

        column_letter = column_cells[0].column_letter
        ws.column_dimensions[column_letter].width = min(
            max(max_length + 3, 12),
            40,
        )


def add_list_validation(ws, column_name, values):
    headers = {
        cell.value: cell.column
        for cell in ws[1]
    }

    column_number = headers.get(column_name)

    if not column_number:
        return

    column_letter = ws.cell(
        row=1,
        column=column_number,
    ).column_letter

    validation = DataValidation(
        type="list",
        formula1='"{}"'.format(",".join(values)),
        allow_blank=False,
    )

    ws.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}1000")


def create_xlsx(
    filename,
    resource_class,
    rows,
    extra_headers=None,
    validations=None,
):
    headers = get_import_headers(resource_class)

    for header in extra_headers or []:
        if header not in headers:
            headers.append(header)

    wb = Workbook()
    ws = wb.active
    ws.title = "IMPORTACAO"

    ws.append(headers)

    for row in rows:
        ws.append([
            row.get(header, "")
            for header in headers
        ])

    apply_header_style(ws)

    for validation in validations or []:
        add_list_validation(
            ws,
            validation["column"],
            validation["values"],
        )

    date_headers = {
        "DATA VENCIMENTO",
    }

    for header in date_headers:
        if header not in headers:
            continue

        column_number = headers.index(header) + 1

        for row_number in range(2, ws.max_row + 1):
            ws.cell(
                row=row_number,
                column=column_number,
            ).number_format = "DD/MM/YYYY"

    text_headers = {
        "CPF/CNPJ",
        "CEP",
        "CONTATO",
    }

    for header in text_headers:
        if header not in headers:
            continue

        column_number = headers.index(header) + 1

        for row_number in range(2, ws.max_row + 1):
            ws.cell(
                row=row_number,
                column=column_number,
            ).number_format = "@"

    adjust_column_widths(ws)

    output_path = OUTPUT_DIR / filename
    wb.save(output_path)

    print(f"Gerado: {output_path}")


def get_carteira_teste():
    carteira = (
        Carteiras
        .objects
        .order_by("id")
        .first()
    )

    if carteira:
        return carteira.nome

    print(
        "AVISO: Nenhuma Carteira foi encontrada no banco. "
        "O arquivo 04 será criado com CARTEIRA TESTE. "
        "Antes de importar, substitua pelo nome de uma Carteira existente."
    )

    return "CARTEIRA TESTE"


def main():
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    carteira = get_carteira_teste()

    # ============================================================
    # 01 - DEVEDORES
    # NOME, CPF/CNPJ
    # ============================================================

    create_xlsx(
        filename="01_devedores.xlsx",
        resource_class=DevedoresResource,
        rows=[
            {
                "NOME": "ANA SOUZA",
                "CPF/CNPJ": "123.456.789-09",
            },
            {
                "NOME": "CARLOS OLIVEIRA",
                "CPF/CNPJ": "987.654.321-00",
            },
            {
                "NOME": "EMPRESA EXEMPLO LTDA",
                "CPF/CNPJ": "12.345.678/0001-95",
            },
        ],
    )

    # ============================================================
    # 02 - ENDERECOS
    # CPF/CNPJ, CEP, LOGRADOURO, BAIRRO, MUNICIPIO, UF
    # ============================================================

    create_xlsx(
        filename="02_enderecos.xlsx",
        resource_class=DevedoresEnderecosResource,
        rows=[
            {
                "CPF/CNPJ": "123.456.789-09",
                "CEP": "01310-100",
                "LOGRADOURO": "AVENIDA PAULISTA, 1000",
                "BAIRRO": "BELA VISTA",
                "MUNICIPIO": "SAO PAULO",
                "UF": "SP",
            },
            {
                "CPF/CNPJ": "123.456.789-09",
                "CEP": "04538-132",
                "LOGRADOURO": "RUA EXEMPLO, 200",
                "BAIRRO": "ITAIM BIBI",
                "MUNICIPIO": "SAO PAULO",
                "UF": "SP",
            },
            {
                "CPF/CNPJ": "987.654.321-00",
                "CEP": "20040-020",
                "LOGRADOURO": "RUA DA ASSEMBLEIA, 50",
                "BAIRRO": "CENTRO",
                "MUNICIPIO": "RIO DE JANEIRO",
                "UF": "RJ",
            },
        ],
        validations=[
            {
                "column": "UF",
                "values": [
                    "AC", "AP", "AM", "PA", "RO", "RR", "TO",
                    "AL", "BA", "CE", "MA", "PB", "PE", "PI",
                    "RN", "SE", "DF", "GO", "MT", "MS", "ES",
                    "MG", "RJ", "SP", "PR", "RS", "SC",
                ],
            },
        ],
    )

    # ============================================================
    # 03 - CONTATOS
    # CPF/CNPJ, TIPO, CONTATO, CONFIANÇA
    # ============================================================

    create_xlsx(
        filename="03_contatos.xlsx",
        resource_class=DevedoresContatosResource,
        rows=[
            {
                "CPF/CNPJ": "123.456.789-09",
                "TIPO": "TELEFONE",
                "CONTATO": "+55 11 99999-9999",
                "CONFIANÇA": "HOT",
            },
            {
                "CPF/CNPJ": "123.456.789-09",
                "TIPO": "EMAIL",
                "CONTATO": "ana.souza@example.com",
                "CONFIANÇA": "HOT",
            },
            {
                "CPF/CNPJ": "987.654.321-00",
                "TIPO": "TELEFONE",
                "CONTATO": "+55 21 98888-7777",
                "CONFIANÇA": "DESCONHECIDO",
            },
            {
                "CPF/CNPJ": "987.654.321-00",
                "TIPO": "EMAIL",
                "CONTATO": "carlos.oliveira@example.com",
                "CONFIANÇA": "VAZIO",
            },
        ],
        validations=[
            {
                "column": "TIPO",
                "values": [
                    "TELEFONE",
                    "EMAIL",
                ],
            },
            {
                "column": "CONFIANÇA",
                "values": [
                    "HOT",
                    "INVALIDO",
                    "DESCONHECIDO",
                    "VAZIO",
                ],
            },
        ],
    )

    # ============================================================
    # 04 - CONTRATOS E PARCELAS
    #
    # The current Resource has CARTEIRA, CPF/CNPJ and PRODUTO as
    # import fields. N° PARCELA, DATA VENCIMENTO and VALOR are
    # consumed manually by after_save_instance(), so they are
    # added here as extra headers.
    # ============================================================

    create_xlsx(
        filename="04_contratos_parcelas.xlsx",
        resource_class=ContratosParcelasResource,
        extra_headers=[
            "N° PARCELA",
            "DATA VENCIMENTO",
            "VALOR",
        ],
        rows=[
            {
                "CARTEIRA": carteira,
                "CPF/CNPJ": "123.456.789-09",
                "PRODUTO": "FINANCIAMENTO",
                "N° PARCELA": 1,
                "DATA VENCIMENTO": date(2026, 9, 10),
                "VALOR": 125000,
            },
            {
                "CARTEIRA": carteira,
                "CPF/CNPJ": "123.456.789-09",
                "PRODUTO": "FINANCIAMENTO",
                "N° PARCELA": 2,
                "DATA VENCIMENTO": date(2026, 10, 10),
                "VALOR": 125000,
            },
            {
                "CARTEIRA": carteira,
                "CPF/CNPJ": "123.456.789-09",
                "PRODUTO": "FINANCIAMENTO",
                "N° PARCELA": 3,
                "DATA VENCIMENTO": date(2026, 11, 10),
                "VALOR": 125000,
            },
            {
                "CARTEIRA": carteira,
                "CPF/CNPJ": "987.654.321-00",
                "PRODUTO": "CREDITO PESSOAL",
                "N° PARCELA": 1,
                "DATA VENCIMENTO": date(2026, 9, 15),
                "VALOR": 85000,
            },
            {
                "CARTEIRA": carteira,
                "CPF/CNPJ": "987.654.321-00",
                "PRODUTO": "CREDITO PESSOAL",
                "N° PARCELA": 2,
                "DATA VENCIMENTO": date(2026, 10, 15),
                "VALOR": 85000,
            },
        ],
    )

    print("")
    print("Arquivos de teste gerados com sucesso.")
    print("")
    print("Ordem sugerida para testar:")
    print("1. 01_devedores.xlsx")
    print("2. 02_enderecos.xlsx")
    print("3. 03_contatos.xlsx")
    print("4. 04_contratos_parcelas.xlsx")


if __name__ == "__main__":
    main()
